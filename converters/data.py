import csv
import json
import xml.etree.ElementTree as ET
from pathlib import Path
from utils.helpers import gerar_nome_saida
from utils.validators import validar_arquivo
from utils.logger import setup_logger

logger = setup_logger(__name__)


class DataConverter:
    """Converte entre CSV, JSON, XML, XLSX, YAML"""

    def __init__(self):
        self.tipo = 'data'
        self.logger = logger

    def _validar(self, arquivo):
        val = validar_arquivo(arquivo, self.tipo)
        if not val['valido']:
            raise ValueError(val['erro'])
        return Path(arquivo)

    # ==================== CSV ====================

    def csv_para_json(self, arquivo):
        """CSV → JSON"""
        arquivo = self._validar(arquivo)

        with open(arquivo, 'r', encoding='utf-8-sig', newline='') as f:
            dados = list(csv.DictReader(f))

        saida = gerar_nome_saida(arquivo, "_converted", ".json")
        with open(saida, 'w', encoding='utf-8') as f:
            json.dump(dados, f, indent=2, ensure_ascii=False)
        return saida

    def csv_para_xml(self, arquivo):
        """CSV → XML"""
        arquivo = self._validar(arquivo)

        root = ET.Element("data")
        with open(arquivo, 'r', encoding='utf-8-sig', newline='') as f:
            for row in csv.DictReader(f):
                item = ET.SubElement(root, "item")
                for key, value in row.items():
                    elem = ET.SubElement(item, _tag_seguro(key))
                    elem.text = value

        saida = gerar_nome_saida(arquivo, "_converted", ".xml")
        ET.ElementTree(root).write(str(saida), encoding='utf-8', xml_declaration=True)
        return saida

    def csv_para_xlsx(self, arquivo):
        """CSV → XLSX"""
        import openpyxl
        arquivo = self._validar(arquivo)

        wb = openpyxl.Workbook()
        ws = wb.active
        with open(arquivo, 'r', encoding='utf-8-sig', newline='') as f:
            for row in csv.reader(f):
                ws.append(row)

        saida = gerar_nome_saida(arquivo, "_converted", ".xlsx")
        wb.save(str(saida))
        return saida

    # ==================== JSON ====================

    def json_para_csv(self, arquivo):
        """JSON → CSV"""
        arquivo = self._validar(arquivo)

        with open(arquivo, 'r', encoding='utf-8') as f:
            dados = json.load(f)

        if not isinstance(dados, list):
            dados = [dados]
        if not dados:
            raise ValueError("JSON vazio")

        # Une as chaves de TODOS os registros (JSON pode ter objetos com
        # campos diferentes entre si; usar só dados[0].keys() quebrava
        # nesse caso).
        fieldnames = []
        for row in dados:
            if isinstance(row, dict):
                for k in row.keys():
                    if k not in fieldnames:
                        fieldnames.append(k)

        saida = gerar_nome_saida(arquivo, "_converted", ".csv")
        with open(saida, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, restval='', extrasaction='ignore')
            writer.writeheader()
            writer.writerows(dados)
        return saida

    def json_para_xml(self, arquivo):
        """JSON → XML"""
        arquivo = self._validar(arquivo)

        with open(arquivo, 'r', encoding='utf-8') as f:
            dados = json.load(f)

        root = ET.Element("data")
        _dict_to_xml(root, dados)

        saida = gerar_nome_saida(arquivo, "_converted", ".xml")
        ET.ElementTree(root).write(str(saida), encoding='utf-8', xml_declaration=True)
        return saida

    def json_para_xlsx(self, arquivo):
        """JSON → XLSX"""
        import openpyxl
        arquivo = self._validar(arquivo)

        with open(arquivo, 'r', encoding='utf-8') as f:
            dados = json.load(f)
        if not isinstance(dados, list):
            dados = [dados]

        wb = openpyxl.Workbook()
        ws = wb.active

        if dados and isinstance(dados[0], dict):
            headers = list(dados[0].keys())
            ws.append(headers)
            for row in dados:
                ws.append([row.get(h, '') for h in headers])

        saida = gerar_nome_saida(arquivo, "_converted", ".xlsx")
        wb.save(str(saida))
        return saida

    def json_para_yaml(self, arquivo):
        """JSON → YAML"""
        import yaml
        arquivo = self._validar(arquivo)

        with open(arquivo, 'r', encoding='utf-8') as f:
            dados = json.load(f)

        saida = gerar_nome_saida(arquivo, "_converted", ".yaml")
        with open(saida, 'w', encoding='utf-8') as f:
            yaml.dump(dados, f, default_flow_style=False, allow_unicode=True, sort_keys=False)
        return saida

    # ==================== XML ====================

    def xml_para_csv(self, arquivo):
        """XML → CSV (espera elementos filhos repetidos tipo <item>/<row>)"""
        arquivo = self._validar(arquivo)
        root = ET.parse(str(arquivo)).getroot()

        candidatos = list(root)
        if not candidatos:
            raise ValueError("XML não contém elementos filhos para converter")

        fieldnames = []
        rows = []
        for item in candidatos:
            row = {child.tag: (child.text or '') for child in item}
            if not row:
                continue
            rows.append(row)
            for k in row:
                if k not in fieldnames:
                    fieldnames.append(k)

        if not rows:
            raise ValueError("Nenhum item com sub-elementos encontrado no XML")

        saida = gerar_nome_saida(arquivo, "_converted", ".csv")
        with open(saida, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, restval='')
            writer.writeheader()
            writer.writerows(rows)
        return saida

    def xml_para_json(self, arquivo):
        """XML → JSON"""
        arquivo = self._validar(arquivo)
        root = ET.parse(str(arquivo)).getroot()

        dados = {root.tag: _xml_to_dict(root)}

        saida = gerar_nome_saida(arquivo, "_converted", ".json")
        with open(saida, 'w', encoding='utf-8') as f:
            json.dump(dados, f, indent=2, ensure_ascii=False)
        return saida

    def xml_para_yaml(self, arquivo):
        """XML → YAML"""
        import yaml
        arquivo = self._validar(arquivo)
        root = ET.parse(str(arquivo)).getroot()

        dados = {root.tag: _xml_to_dict(root)}

        saida = gerar_nome_saida(arquivo, "_converted", ".yaml")
        with open(saida, 'w', encoding='utf-8') as f:
            yaml.dump(dados, f, default_flow_style=False, allow_unicode=True, sort_keys=False)
        return saida

    # ==================== XLSX ====================

    def xlsx_para_csv(self, arquivo):
        """XLSX → CSV (primeira planilha)"""
        import openpyxl
        arquivo = self._validar(arquivo)

        wb = openpyxl.load_workbook(str(arquivo), data_only=True)
        ws = wb.active

        saida = gerar_nome_saida(arquivo, "_converted", ".csv")
        with open(saida, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(['' if v is None else v for v in row])
        return saida

    def xlsx_para_json(self, arquivo):
        """XLSX → JSON (primeira linha = cabeçalho)"""
        import openpyxl
        arquivo = self._validar(arquivo)

        wb = openpyxl.load_workbook(str(arquivo), data_only=True)
        ws = wb.active

        linhas = list(ws.iter_rows(values_only=True))
        if not linhas:
            raise ValueError("Planilha vazia")

        headers = [str(h) if h is not None else '' for h in linhas[0]]
        dados = [dict(zip(headers, row)) for row in linhas[1:]]

        saida = gerar_nome_saida(arquivo, "_converted", ".json")
        with open(saida, 'w', encoding='utf-8') as f:
            json.dump(dados, f, indent=2, ensure_ascii=False, default=str)
        return saida

    # ==================== YAML ====================

    def yaml_para_json(self, arquivo):
        """YAML → JSON"""
        import yaml
        arquivo = self._validar(arquivo)

        with open(arquivo, 'r', encoding='utf-8') as f:
            dados = yaml.safe_load(f)

        saida = gerar_nome_saida(arquivo, "_converted", ".json")
        with open(saida, 'w', encoding='utf-8') as f:
            json.dump(dados, f, indent=2, ensure_ascii=False)
        return saida

    def yaml_para_xml(self, arquivo):
        """YAML → XML"""
        import yaml
        arquivo = self._validar(arquivo)

        with open(arquivo, 'r', encoding='utf-8') as f:
            dados = yaml.safe_load(f)

        root = ET.Element("data")
        _dict_to_xml(root, dados)

        saida = gerar_nome_saida(arquivo, "_converted", ".xml")
        ET.ElementTree(root).write(str(saida), encoding='utf-8', xml_declaration=True)
        return saida


# ==================== Helpers de módulo ====================

def _tag_seguro(nome):
    """Garante que uma string vira uma tag XML válida (não pode começar com número, etc.)"""
    nome = str(nome).strip().replace(' ', '_')
    if not nome or not (nome[0].isalpha() or nome[0] == '_'):
        nome = f"campo_{nome}" if nome else "campo"
    return nome


def _dict_to_xml(parent, data):
    if isinstance(data, dict):
        for key, value in data.items():
            elem = ET.SubElement(parent, _tag_seguro(key))
            _dict_to_xml(elem, value)
    elif isinstance(data, list):
        for item in data:
            elem = ET.SubElement(parent, "item")
            _dict_to_xml(elem, item)
    else:
        parent.text = "" if data is None else str(data)


def _xml_to_dict(elem):
    if len(elem) == 0:
        return elem.text or ''
    result = {}
    for child in elem:
        child_data = _xml_to_dict(child)
        if child.tag in result:
            if not isinstance(result[child.tag], list):
                result[child.tag] = [result[child.tag]]
            result[child.tag].append(child_data)
        else:
            result[child.tag] = child_data
    return result
